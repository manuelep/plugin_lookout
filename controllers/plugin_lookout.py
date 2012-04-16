
# coding: utf8

#    This file is part of plugin_lookout.

#    Plugin_lookout is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    Plugin_lookout is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with Plugin_lookout.  If not, see <http://www.gnu.org/licenses/>.

#from random import randint

def index(): return dict()

#        # condizione: esiste tabella con lo stesso nome in una connessione con stesso alias
#        alias = db(db.plugin_lookout_connections.id==form.vars.connection)\
#            .select(db.plugin_lookout_connections.alias).first().alias
#        subquery = db(db.plugin_lookout_connections.alias==alias)\
#            ._select(db.plugin_lookout_connections.id)
#        res = db(db.plugin_lookout_tables.connection.belongs(subquery)\
#            &(db.plugin_lookout_tables.table_name==form.vars.table_name)).select()
#        if len(res)==1:
#            set_table_restrictions(res.first())
#            db.commit()

################################################### PLUGIN_LOOKOUT_CONNECTIONS #

def connection_onvalidation(form):
    '''
    The callback function for the connections form
    '''
    if form.vars.pwd:
        uri = form.vars.dsn % form.vars.pwd
    else:
        uri = form.vars.dsn
    
    if db._uri == uri:
        form.vars.alias = 'db'
    
    res = db(db.plugin_lookout_connections.dsn==form.vars.dsn).select()

    if len(res)>0:
        rec = res.first()
        form.vars.alias = rec.alias
        group_id = auth.user_group(auth.user_id)

        if not auth.has_permission(table_name='plugin_lookout_connections',
            user_id=auth.user_id,
            record_id=rec.id
        ):
            auth.add_permission(table_name='plugin_lookout_connections',
                group_id=group_id, record_id=rec.id)

        db_hash = globals()[form.vars.alias]._uri_hash
        groups = db(db.auth_group.role.contains(db_hash)).select(db.auth_group.id)
        for group in groups:
            if not auth.has_membership(group.id, auth.user_id):
                auth.add_membership(group.id, auth.user_id)

def connection_oncreate(form):
    set_data_permission('plugin_lookout_connections', form.vars.id)

@auth.requires_login()
def plugin_lookout_connections():
    
#    db.plugin_lookout_connections.dsn.represent = lambda v, r: '%s: %s' % (r.alias, v.replace('%s', randint(8, 16)*'*'))
    form = SQLFORM.smartgrid(db.plugin_lookout_connections,
        onvalidation = connection_onvalidation,
        linked_tables = [],
        constraints = dict(plugin_lookout_connections=get_connection_set())
    )
    
    return dict(form=form)

######################################################## PLUGIN_LOOKOUT_TABLES #

def table_onvalidate(form):
    '''
    The callback function for the tables form
    '''
    # before creation
    from plugin_lookout import db_got_table
    if 'new' in request.args:
        connection_name = db.plugin_lookout_connections[form.vars.connection_id].alias
        mydb = globals()[connection_name]
        tab_in_sqldb, msg = db_got_table(mydb, form.vars.table_name)
        
        if tab_in_sqldb:
            form.vars.table_migrate = False
        else:
            form.vars.table_migrate = True

        form.errors.table_name = IS_VALID_SQL_TABLE_NAME(mydb,
            check_reserved = ('common', 'postgres', )
        )(form.vars.table_name)[1]
    
    # before deletion
    if form.vars.delete_this_record=='on':
#        join = db.plugin_lookout_tables.connection_id==db.plugin_lookout_connections.id
        where = db.plugin_lookout_tables.id==request.vars.id
        rec_table = db(where).select(
            db.plugin_lookout_tables.table_name,
            db.plugin_lookout_tables.table_migrate,
            db.plugin_lookout_tables.is_view,
            db.plugin_lookout_tables.is_active,
            db.plugin_lookout_tables.connection_name
        ).first()
        table_name = rec_table.table_name
        table_migrate = rec_table.table_migrate
        table_is_view = rec_table.is_view
        connection_name = rec_table.connection_name
        table_is_active = rec_table.is_active
        
        key = '%s_%s' % (globals()[connection_name]._uri_hash, table_name)
        db(db.auth_group.role.contains(key)).delete()
        
        db(db.plugin_lookout_fields.table_id==request.vars.id).delete()
        
#        for row in db(db.plugin_lookout_fields.tables.contains(form.vars.id)).select():
#            ids = row.tables
#            ids.remove(int(form.vars.id))
#            row.update_record(tables=ids)
            
        if not table_migrate and table_is_view:
            db.executesql('DROP VIEW %s CASCADE;' % table_name)
        elif table_migrate and table_is_active:
            try:
                globals()[connection_name][table_name].drop()
            except Exception, error:
                msg = T('Table %s not removed: %s') % (table_name, str(error))
                session.flash = msg
                # this will block record deletion
                form.errors.external_error = str(msg)
        
def table_oncreate(form):
    # setupo permission only on creation and on update
    if form.vars.delete_this_record != 'on':
        if form.vars.is_active:
            if 'connection_id' in form.vars:
                connection_name = db.plugin_lookout_connections[form.vars.connection_id].alias
            else:
                connection_name = db.plugin_lookout_tables[form.vars.id].connection_id
            if form.vars.restricted:
                t_hash = '%s_%s' % (globals()[connection_name]._uri_hash, form.vars.table_name)
                w_role = 'write_%s' % t_hash
                r_role = 'read_%s' % t_hash
                set_data_permission('plugin_lookout_tables', form.vars.id, role=r_role)
                set_data_permission('plugin_lookout_tables', form.vars.id, role=w_role)
        
def table_ondelete():
    unusefull_fields_set = db(db.plugin_lookout_fields.tables==[])
    if unusefull_fields_set.count() > 0:
        unusefull_fields_set.delete()
      
@auth.requires_login()
def plugin_lookout_tables():
    '''
    Controller for managing table features and settings.
    funzioni da aggiungere:
    * aggiungi campo a tabella
    * rimuovi campo dalla tabella
    '''
    
    if 'edit' in request.args:
        db.plugin_lookout_tables.table_name.writable = False
        db.plugin_lookout_tables.connection_id.writable = False

    only_view = bool(request.vars.only_view)
    session.plugin_lookout_tables_only_view = only_view
    if 'edit' in request.vars:
        editable = auth.has_permission('plugin_lookout_tables',
            request.args(0),
            auth.user_id
        )
    else:
        editable = not session.plugin_lookout_tables_only_view

    table_set = get_table_set(view_only=bool(only_view))

    db.plugin_lookout_tables.table_name.represent = lambda val,row: A(row.table_name, _href=URL('plugin_lookout_external_tables', vars=dict(table_id=row.id)))
#    db.plugin_lookout_tables.connection_id.represent = lambda val,row: ('%(dsn)s' % db.plugin_lookout_connections[val]).replace('%s', randint(8, 16)*'*')
    form = SQLFORM.smartgrid(db.plugin_lookout_tables,
        onvalidation = table_onvalidate,
        oncreate = table_oncreate,
        onupdate = table_oncreate,
        ondelete = table_ondelete,
        deletable = ('edit' in request.args),
        editable = editable,
        linked_tables = [],
        constraints = dict(plugin_lookout_tables=table_set) # get_data_with_permissions('plugin_lookout_tables')
    )
    
    plugin_lookout_tables_menu = None
    if 'view' in request.args:
        table_id = request.args[-1]
        plugin_lookout_tables_menu = [
            (T('Share data'), False, URL('share_data_with_users', args=table_id), []),
            (T('Edit table fields'), False, URL('plugin_lookout_fields',
                vars=dict(keywords='plugin_lookout_fields.table_id="%s"' % table_id)), [])
        ]
        if db(db.plugin_lookout_datafiles.table_id==table_id).count():
            plugin_lookout_tables_menu += [
                (T('Init/Reset table data'), False, URL('init_external_table', vars=dict(table_id=table_id)), []),
            ]
    return dict(form=form, plugin_lookout_tables_menu=plugin_lookout_tables_menu)

######################################################## PLUGIN_LOOKOUT_FIELDS #

#db.plugin_lookout_fields.table.represent = lambda id,row: CAT(*[CAT(A('%s ' %i.table_name,
#    _href = URL('plugin_lookout_tables',
#        args = ['plugin_lookout_tables', 'view','plugin_lookout_tables', i.id],
#        user_signature = True)), BR()
#    ) for i in db(db.plugin_lookout_tables.id.belongs(id))\
#        .select(db.plugin_lookout_tables.id, db.plugin_lookout_tables.table_name)])

@auth.requires(control_permission(request.vars.table_id, reading=False, default=True), requires_login=True)
def plugin_lookout_fields():
    '''
    Controller for managing field features and settings.
    '''

    table_condition = get_table_set(view_only=False)
    db.plugin_lookout_fields.table_id.requires = IS_IN_DB(db(get_table_set(view_only=False)), db.plugin_lookout_tables.id, '%(table_name)s')
    db.plugin_lookout_fields.table_id.represent = lambda value, row: db.plugin_lookout_tables[value].table_name
    form = SQLFORM.smartgrid(db.plugin_lookout_fields,
        linked_tables=['plugin_lookout_tables'],
        orderby = db.plugin_lookout_fields.table_id,
#        editable=False,
#        deletable=False,
        constraints = dict(plugin_lookout_fields=(db.plugin_lookout_fields\
            .table_id.belongs(db(table_condition)._select(db.plugin_lookout_tables.id))))
    )
    return dict(form=form)

####################################################################### TABLES #

@auth.requires(control_permission(request.vars.table_id or session.plugin_lookout_external_tables_id, reading=True), requires_login=True)
def plugin_lookout_external_tables():
    '''
    Controller for manage data inside table that are not part of the model.
    It is called from the linked table name in the plugin_lookout_tables edit grid.
    '''
    message = 'Here you can see the date inside the tables you have configured'

    table_id = request.vars.table_id or session.plugin_lookout_external_tables_id or redirect(URL('plugin_lookout_tables'))
    session.plugin_lookout_external_tables_id = table_id
    
    check_message = IS_IN_DB(db, 'plugin_lookout_tables.id')(table_id)[1]
    if check_message:
#        session.flash = 'Non puoi accedere alla tabella "%s" attraverso questa risorsa. %s' %(tab_name, check_message)
        session.flash = T('You cannot have access to the table "%s" through this resource. %s' %(table_id, check_message))
        redirect(URL('plugin_lookout_tables'))
    
    rec_table = db.plugin_lookout_tables[table_id]
    mydb = globals()[rec_table.connection_name]
    if rec_table.table_name not in mydb.tables:
        session.flash = T('Table "%s" is not recognized from db model. maybe it\'s not active') % rec_table.table_name
        session.flash = 'La tabella "%s" non riconosciuta in database o non attiva.' % rec_table.table_name
        redirect(URL('plugin_lookout_tables'))
    
    writable=control_permission(table_id, reading=False)
    grid=SQLFORM.smartgrid(mydb[rec_table.table_name], deletable=writable, editable=writable, create=writable)
    if request.extension == 'load':
        return dict(grid=grid)
    else:
        return dict(grid=grid, message=message)

################################################################### SHARE DATA #

def group_representation(value, row):
    rec_group = db.auth_group[value]
    representation = rec_group.role
    import ipdb; ipdb.set_trace()
    if rec_group.role[:5] == 'user_':
        user_id = int(rec_group.role.split('_')[1])
        rec_user = db.auth_user[user_id]
        representation += ': %(first_name)s %(last_name)s'
        return representation % rec_user
    else:
        return representation

@auth.requires_login()
def share_data_with_users():
    '''
    This controller is for choose the user groups that you want to share your
    data with. It is called from the plugin_lookout_tables edit grid context menu.
    '''
    table_id = request.args(0) or redirect(URL('plugin_lookout_tables'))
#    join = (db.auth_membership.group_id==db.auth_group.id)\
#        &(db.auth_membership.user_id==db.auth_user.id)
    form = SQLFORM.factory(
        Field('groups', 'list:reference auth_group',
            requires=IS_IN_DB(db, 'auth_group.id', '%(role)s', multiple=True),
        ),
        Field('read_only', 'boolean')
    )
    if form.accepts(request, session):
        rec_table = db.plugin_lookout_tables[table_id]
        tab_db = globals()[rec_table.connection_name]
        t_hash = '%s_%s' % (tab_db._uri_hash, rec_table.table_name)
        if form.vars.read_only:
            role = 'read_%s' % t_hash
        else:
            role = 'write_%s' % t_hash
        
        for group_id in form.vars.groups:
            set_data_permission('plugin_lookout_tables', table_id, role=role, group_id=group_id)
    
    
#        groups = db(db.auth_user.id.belongs(form.vars.groups)).select()
#        join = db.plugin_lookout_tables.connection==db.plugin_lookout_connections.id
#        rec_table = db(join&(db.plugin_lookout_tables.id==table_id)).select(
#            db.plugin_lookout_tables.table_name,
#            db.plugin_lookout_connections.alias
#        ).first()
#        table = globals()[rec_table.plugin_lookout_connections.alias][rec_table.plugin_lookout_tables.table_name]
#        share_data(table, read_only=form.vars.read_only, users=users)
    return dict(form=form)
    

################################################################### XLS IMPORT #

from openpyxl.reader.excel import load_workbook
from plugin_lookout import guess_type

@auth.requires_login()
def import_xls_structure():
    '''
    This controller is for building a table structure and contain data from an
    excell spread sheet. It creates as many fields as the number of the columns
    if finds in the first sheet of the xls file and try to guess the type of
    data to contain for each field. The table structure is created not active so
    you can made some change before to import data.
    After the table creation you'll be redirected to the field managment
    controller filtered on the fields of the newly created table.
    It is called from the plugin_lookout_tables edit grid context menu.
    
    TODO: In the future all sheets of the xls file can be supported and more
    than one table structure can be created in one time. Maybe could be asked to
    the user how many fields to import (0 for all sheets)
    '''

    db.plugin_lookout_tables.is_active.default = False
    db.plugin_lookout_tables.is_active.writable = False
    if not 'new' in request.args: redirect(URL('plugin_lookout_tables'))

    db.plugin_lookout_datafiles.extension.requires = plugin_lookout_datafiles_types[1:2]
    form = SQLFORM.factory(
        db.plugin_lookout_tables,
        db.plugin_lookout_datafiles
#        Field('source_file', 'upload', uploadfolder=uploadfolder)
    )

    if form.accepts(request, session, onvalidation=table_onvalidate):
    
        table_id = db.plugin_lookout_tables.insert(**db.plugin_lookout_tables._filter_fields(form.vars))
        form.vars.table_id=table_id
        file_id = db.plugin_lookout_datafiles.insert(**db.plugin_lookout_datafiles._filter_fields(form.vars))
    
        filePath = os.path.join(uploadfolder, db.plugin_lookout_datafiles[file_id].file_name)
        
        xlsx = load_workbook(filePath)
        first_sheet_name = xlsx.get_sheet_names()[0]
        sheet = xlsx.get_sheet_by_name(first_sheet_name)
        header = sheet.rows[0]

        for idx, cell in enumerate(header):
            values = [i.value for i in sheet.columns[idx]][1:]
            
            field_type = guess_type(values)
            
            ret = db.plugin_lookout_fields.validate_and_insert(
                table_id = table_id,
                field_name = cell.value.lower().replace(' ', ''),
                field_comment = cell.value,
                field_type = field_type
            )
            if ret.get('error'):
                db.plugin_lookout_fields.insert(
                    table_id = table_id,
                    field_name = 'field_%s' % idx,
                    field_label = cell.value,
                    field_comment = str(ret.error),
                    field_type = field_type
                )
        xlsx.close()
        redirect(URL('plugin_lookout_fields',
                vars=dict(keywords='plugin_lookout_fields.table_id="%s"' % table_id)))
    
    return dict(form=form)

@auth.requires_login()
def init_xls_data():
    '''
    Once a table structure is builded starting from an uploaded xls file, once
    you've verified that all data type correnspond to the data you're bound to
    upload than you can use this controller to upload data from the same file
    stored in data base. I this way the new table can be resetted to the starting
    condition whenever you need. The stored file can be considered as a data backup.
    '''
    table_id = request.vars.table_id or redirect(URL('plugin_lookout_tables'))
    
    table_info = db.plugin_lookout_tables[table_id]
    
    if not table_info.is_active:
        session.flash = T('You cannot populate not active tables')
        redirect(URL('plugin_lookout_tables', args=['plugin_lookout_tables', 'edit', 'plugin_lookout_tables', table_id]))
    
    filePath = os.path.join(uploadfolder, db.plugin_lookout_datafiles[table_id].file_name)
    xlsx = load_workbook(filePath)
    first_sheet_name = xlsx.get_sheet_names()[0]
    sheet = xlsx.get_sheet_by_name(first_sheet_name)
    
    res_fields = db(db.plugin_lookout_fields.table_id==table_id).select(db.plugin_lookout_fields.field_name)
    
    field_names = [i.field_name for i in res_fields]
    
    mydb = globals()[table_info.connection_name]
    
    if mydb(mydb[table_info.table_name]).count(): mydb[table_info.table_name].drop()
    
    error = None
    for row in sheet.rows[1:]:
        values = [cell.value for cell in row]
        
        kwargs = dict([(k,v) for k,v in zip(field_names, values)])
        ret = mydb[table_info.table_name].validate_and_insert(**kwargs)
        if ret.errors:
            error = T('Import error in line %(line)s, column %(key)s: %(value)s (%(data)s)', dict(line=n, data=data, **ret.errors))
            mydb.rollback()
            break
    
    if error:
        session.flash = error
        redirect(URL('plugin_lookout_tables'))
    
    redirect(URL('plugin_lookout_external_tables', vars=dict(table_id=table_id)))


################################################################# IMPORT SHAPE #

from plugin_lookout import file2struct

@auth.requires_login()
def import_struct():
    db.plugin_lookout_tables.is_active.default = True

    db.plugin_lookout_datafiles.extension.requires = plugin_lookout_datafiles_types[1:9]
    form = SQLFORM.factory(
        db.plugin_lookout_tables,
        db.plugin_lookout_datafiles
    )

    if form.accepts(request, session, onvalidation=table_onvalidate):
        table_id = db.plugin_lookout_tables.insert(**db.plugin_lookout_tables._filter_fields(form.vars))
        form.vars.table_id=table_id
        file_id = db.plugin_lookout_datafiles.insert(**db.plugin_lookout_datafiles._filter_fields(form.vars))
        file_path = os.path.join(uploadfolder, db.plugin_lookout_datafiles[file_id].file_name)
        file_name = db.plugin_lookout_datafiles[file_id].file_name
        fileExt = fileName.split('.')[-1]
        
        file2struct(file_name, uploadfolder, table_id, db.plugin_lookout_fields)
        
        redirect(URL('plugin_lookout_fields',
            vars=dict(keywords='plugin_lookout_fields.table_id="%s"' % table_id)))
        
    return dict(form=form)

from plugin_lookout import initFromFile

@auth.requires_login()
def init_external_table():
    table_id = request.vars.table_id or redirect(URL('plugin_lookout_tables'))
    
    table_info = db.plugin_lookout_tables[table_id]
    
    if not table_info.is_active:
        session.flash = T('You cannot populate not active tables')
        redirect(URL('plugin_lookout_tables', args=['plugin_lookout_tables', 'edit', 'plugin_lookout_tables', table_id]))
    
    ext_table = globals()[table_info.connection_name][table_info.table_name]
    
    initFromFile(db(db.plugin_lookout_datafiles.table_id==table_id).select().first().file_name, uploadfolder, table_id, db, ext_table)
    
    redirect(URL('plugin_lookout_external_tables', vars=dict(table_id=table_id)))
    
    










#@auth.requires_login()
#def import_shp():
#    '''
#    '''
#    from archive import extract
#    import ogr, shutil
#    db.plugin_lookout_tables.is_active.default = True
#    db.plugin_lookout_tables.is_active.writable = False
##    if not 'new' in request.args: redirect(URL('plugin_lookout_tables'))

#    db.plugin_lookout_datafiles.extension.requires = plugin_lookout_datafiles_types[1:9]
#    form = SQLFORM.factory(
#        db.plugin_lookout_tables,
#        db.plugin_lookout_datafiles
#    )
#    
#    ids=None
#    if form.accepts(request, session, onvalidation=table_onvalidate):
#    
#        table_id = db.plugin_lookout_tables.insert(**db.plugin_lookout_tables._filter_fields(form.vars))
#        form.vars.table_id=table_id
#        file_id = db.plugin_lookout_datafiles.insert(**db.plugin_lookout_datafiles._filter_fields(form.vars))
#        file_path = os.path.join(uploadfolder, db.plugin_lookout_datafiles[file_id].file_name)
#        
#        # uncompress the archive in new_dir
#        new_dir = '.'.join(form.vars.file_name.split('.')[:-1])
#        new_path = os.path.join(uploadfolder, new_dir)
#        os.mkdir(new_path)
#        ext = file_path.split('.')[-1]
#        if ext in ('gz', 'bz2', ):
#            file_path_new = '.'.join(file_path.split('.')[:-1] + ['tar', ext])
#            os.rename(file_path, file_path_new)
#            file_path = file_path_new
#        extract(file_path, new_path)
#        main_shp = [i for i in os.listdir(new_path) if i.split('.')[-1]=='shp'][0]
#        shp_path = os.path.join(new_path, main_shp)
#        
#        driver = ogr.GetDriverByName('ESRI Shapefile')
#        source = driver.Open(shp_path, 0)
#        layer = source.GetLayer()
#        # inspect field names and types
#        ESRITypes = dict(String='string', Real='double', Date='date')
#        layer_defn = layer.GetLayerDefn()
#        layer_infos = [(layer_defn.GetFieldDefn(i).GetName(),
#            ESRITypes[layer_defn.GetFieldDefn(i).GetTypeName()]) for i in xrange(layer_defn.GetFieldCount())]
#        
#        # setup geometry field
#        ret = db.plugin_lookout_fields.validate_and_insert(
#                table_id = table_id,
#                field_name = 'the_geom',
#                field_label = 'Geometric feature',
#                field_type = 'geometry'
#            )
#        
#        # setup attributes fields
#        for field_name, field_type in layer_infos:
#            ret = db.plugin_lookout_fields.validate_and_insert(
#                table_id = table_id,
#                field_name = field_name.lower(),
#                field_label = field_name,
#                field_type = field_type
#            )
#        
#        import ipdb; ipdb.set_trace()
#        table = globals()[db.plugin_lookout_connections[form.vars.connection_id].alias][form.vars.table_name]
#        for index in xrange(layer.GetFeatureCount()):
#            feature = layer.GetFeature(index)
#            kwargs = dict([(fn[0].lower(), feature.GetField(fn[0])) for fn in layer_infos])
#            if not hasattr(table['the_geom'], 'st_asgeojson'):
#                kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkb()
#            else:
#                kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkt()
#            
#            ret = table.validate_and_insert(**kwargs)
#            if ret.errors:
#                raise IOError(str(ret.errors))
#            else:
#                ids.append = ret.id
#        
#        shutil.rmtree(new_path) # remove uncompressed files
#        redirect(URL('plugin_lookout_fields',
#                vars=dict(keywords='plugin_lookout_fields.table_id="%s"' % table_id)))

#    return dict(form=form, ids=ids)

#def init_shp_data():
#    table_id = request.vars.table_id or redirect(URL('plugin_lookout_tables'))
#    
#    table_info = db.plugin_lookout_tables[table_id]
#    
#    if not table_info.is_active:
#        session.flash = T('You cannot populate not active tables')
#        redirect(URL('plugin_lookout_tables', args=['plugin_lookout_tables', 'edit', 'plugin_lookout_tables', table_id]))
        

#def import_shape():
#    from archive import extract
#    import ogr, shutil
#    form = SQLFORM.factory(
#        Field('connection_name', requires=IS_IN_DB(db(get_connection_set()), 'plugin_lookout_connections.alias', '%(alias)s')),
#        Field('table_name'),
#        Field('shp_archive', 'upload', uploadfolder=uploadfolder, label='File',
#            comment=T('Supported archives are: .zip, .egg, .jar, .tar, .tar.gz, .tgz, .tar.bz2, .tz2'))
#    )
#    
#    ESRITypes = dict(String='string', Real='double', Date='date')
#    if form.accepts(request, session):
#        
#        file_path = os.path.join(uploadfolder, form.vars.shp_archive)
#        new_dir = '.'.join(form.vars.shp_archive.split('.')[:-1])
#        new_path = os.path.join(uploadfolder, new_dir)
#        os.mkdir(new_path)
#        ext = file_path.split('.')[-1]
#        if ext in ('gz', 'bz2', ):
#            file_path_new = '.'.join(file_path.split('.')[:-1] + ['tar', ext])
#            os.rename(file_path, file_path_new)
#            file_path = file_path_new
#        extract(file_path, new_path)
#        main_shp = [i for i in os.listdir(new_path) if i.split('.')[-1]=='shp'][0]
#        form.vars.shp_path = os.path.join(new_path, main_shp)
#        
#        
#        shutil.rmtree(new_dir)
#        
#        driver = ogr.GetDriverByName('ESRI Shapefile')
#        source = driver.Open(form.vars.shp_path, 0)
#        layer = source.GetLayer()
#        layer_defn = layer.GetLayerDefn()
#        field_infos = [(layer_defn.GetFieldDefn(i).GetName(),
#            ESRITypes[layer_defn.GetFieldDefn(i).GetTypeName()]) for i in xrange(layer_defn.GetFieldCount())]
#        
##        for field_name, field_type in field_infos:
##            kwargs = dict(
##                
##            )
##            db.plugin_lookout_fields.
#        
#        for index in range(layer.GetFeatureCount())[:5]:
#            feature = layer.GetFeature(index)
#            print dict([(fn[0], (fn[1], feature.GetField(fn[0]))) for fn in field_names])
#            wkt = feature.GetGeometryRef().ExportToWkt()
#            print wkt
#        

#    return dict(form=form)


################################################################## CREATE VIEW #

def create_view_step1_onvalidation(form):
    ids = [form.vars.table_id] + form.vars.table_ids
    res_tables = db(db.plugin_lookout_tables.id.belongs(ids)).select()
    if len(set([i.connection_name for i in res_tables]))>1:
        error_message = T('Table to join must belong to the same connection')
        form.errors.table_ids = error_message
        form.errors.table_id = error_message
        

@auth.requires_login()
def create_view_step1():
    message = 'Here you can create table views that join two tables with a field in common. \
    Choose the two tables from the lists below.'
    table_set = db(get_table_set(view_only=False))
    form = SQLFORM.factory(
#        Field('view_name', label=T("Views's name"), comment=T("Name of the view to be create"),
#            requires=plugin_metadb.IS_VALID_SQL_TABLE_NAME(globals()[db.plugin_lookout_connections[request.vars.connection].connection_name], check_reserved=('common', 'postgres', ))),
        Field('table_id', 'integer',
            requires = IS_IN_DB(table_set, 'plugin_lookout_tables.id', '%(table_name)s (from %(connection_name)s)')),
        Field('table_ids', 'list:integers',
            requires = IS_IN_DB(table_set, 'plugin_lookout_tables.id', '%(table_name)s (from %(connection_name)s)', multiple=True)
        )
    )
    
    if form.accepts(request, session, onvalidation=create_view_step1_onvalidation):
        try:
            idx = form.vars.table_ids.index(form.vars.table_id)
        except ValueError, error:
            pass
        else:
            form.vars.table_ids.pop(idx)
        print form.vars.table_ids
        if not form.vars.table_ids: return dict(form=form)
        session.create_view_step2 = form.vars
        redirect(URL('create_view_step2'))
    
    return dict(form=form)

@auth.requires_login()
def create_view_step2():
    
    ids = [session.create_view_step2.table_id] + session.create_view_step2.table_ids
    field_set = db(db.plugin_lookout_fields.table_id.belongs(ids))
    form = SQLFORM.factory(
        Field('ext_ref', label=T('Join column name'),
            requires = IS_IN_DB(field_set, 'plugin_lookout_fields.id', '%(field_name)s in table: %(table_id)s')
        )
    )

    return dict(form=form)
