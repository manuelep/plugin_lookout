
#!/usr/bin/env python
# coding: utf8

#    This file is part of plugin_lookout.

#    Plugin_lookout is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.

#    Plugin_lookout is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.

#    You should have received a copy of the GNU General Public License
#    along with Plugin_lookout.  If not, see <http://www.gnu.org/licenses/>.

from gluon import *

from gluon.custom_import import track_changes; track_changes(True)

from dal import regex_python_keywords
from validators import Validator, translate
import re

class IS_VALID_SQL_TABLE_NAME(Validator):
    '''
    Checks if the field value is a valid SQL table name.
    
    Arguments:
    
    db: an instance of the dal class that represent the database that will
        contain the table name
    
    check_reserved: list of adapters to check tablenames and column names
         against sql reserved keywords. (Default ("common", ))
        * "common" List of sql keywords that are common to all database types
            such as "SELECT, INSERT". (recommended)
        * "all" Checks against all known SQL keywords. (not recommended)
            <adaptername> Checks against the specific adapters list of keywords
            (recommended)
    
    Examples:
    
        #Check if text string is a good sql table name:
        INPUT(_type="text", _name="name", requires=IS_VALID_SQL_TABLE_NAME(db))
        
        #Check if text string is a good sql table name specific for postgres dialect:
        INPUT(_type="text", _name="name", requires=IS_VALID_SQL_TABLE_NAME(db, check_reserved=("postgres", )))
        
        >>> IS_VALID_SQL_TABLE_NAME(db)("")
        ('', 'invalid table name: ')
        >>> IS_VALID_SQL_TABLE_NAME(db)("foo")
        ('foo', None)
        >>> IS_VALID_SQL_TABLE_NAME(db)("test")
        ('test', 'table/attribute name already defined: test')
        >>> IS_VALID_SQL_TABLE_NAME(db)("select")
        ('select', 'invalid table/column name "select" is a "COMMON" reserved SQL keyword')
    
    '''
    def __init__(self, db, check_reserved=('common', )):
        self.db = db
        self.check_reserved = check_reserved
        if self.check_reserved:
            from reserved_sql_keywords import ADAPTERS as RSK
            self.RSK = RSK

    def __call__(self, value):
    
        if re.compile('[^0-9a-zA-Z_]').findall(value):
            return (value, translate('only [0-9a-zA-Z_] allowed in table and field names, received %s' % value))
        elif value.startswith('_') or regex_python_keywords.match(value) or not value:
            return (value, translate('invalid table name: %s' % value))
        elif value.lower() in self.db.tables or hasattr(self.db,value.lower()):
            return (value, translate('table/attribute name already defined: %s' % value))
        elif self.check_reserved:
            # Validates ``name`` against SQL keywords
            #+ Uses self.check_reserve which is a list of
            #+ operators to use.
            #+ self.check_reserved
            #+ ['common', 'postgres', 'mysql']
            #+ self.check_reserved
            #+ ['all']
            for backend in self.check_reserved:
                if value.upper() in self.RSK[backend]:
                    return (value, translate('invalid table/column name "%s" is a "%s" reserved SQL keyword' % (value, backend.upper())))
            return (value, None)

def db_got_table(db, table):
    '''
    db: the database connection where to look for
    table: the table name to look for
    '''
#    db_type = db._uri[:8]
    msg = ''
    sql_src = "SELECT 1 FROM %s WHERE 1=2" % table
    try:
        db.executesql(sql_src)
    except Exception, error:
        msg = str(error).split('\n')[0].strip()
        answare = False
    else:
        answare = True
    return answare, msg

def guess_type(values):
    types = map(type, values)
    if float in types:
        return 'double'
    elif int in types:
        return 'integer'
    else:
        return None


def geom_representation(value, n=7):
    if not value: return value
    try:
        import ppygis
    except:
        return value[:n]
    else:
        return str(ppygis.Geometry.read_ewkb(value))[:n]

from openpyxl.reader.excel import load_workbook
from archive import extract
import ogr, os
def file2struct(fileName, path, table_id, plugin_lookout_fields):
    fileExt = fileName.split('.')[-1]
    filePath = os.path.join(path, fileName)
    if fileExt in ('xls', 'xlsx', ):
        xlsx = load_workbook(filePath)
        first_sheet_name = xlsx.get_sheet_names()[0]
        sheet = xlsx.get_sheet_by_name(first_sheet_name)
        header = sheet.rows[0]

        for idx, cell in enumerate(header):
            values = [i.value for i in sheet.columns[idx]][1:]
            
            field_type = guess_type(values)
            
            ret = plugin_lookout_fields.validate_and_insert(
                table_id = table_id,
                field_name = cell.value.lower().replace(' ', ''),
                field_comment = cell.value,
                field_type = field_type
            )
            if ret.get('error'):
                plugin_lookout_fields.insert(
                    table_id = table_id,
                    field_name = 'field_%s' % idx,
                    field_label = cell.value,
                    field_comment = str(ret.error),
                    field_type = field_type
                )
    elif fileExt in ('zip', 'egg', 'jar', 'tar', 'gz', 'tgz', 'bz2', 'tz2', ):
        
        # uncompress the archive in new_dir
        new_dir = '.'.join(fileName.split('.')[:-1])
        new_path = os.path.join(path, new_dir)
        try:
            os.mkdir(new_path)
        except OSError, error:
            pass
        else:
            if fileExt in ('gz', 'bz2', ):
                filePath_new = '.'.join(filePath.split('.')[:-1] + ['tar', fileExt])
                os.rename(filePath, filePath_new)
                filePath = filePath_new
            extract(filePath, new_path)

        main_shp = [i for i in os.listdir(new_path) if i.split('.')[-1]=='shp'][0]
        shp_path = os.path.join(new_path, main_shp)
        
        driver = ogr.GetDriverByName('ESRI Shapefile')
        source = driver.Open(shp_path, 0)
        layer = source.GetLayer()

        # inspect field names and types
        ESRITypes = dict(String='string', Real='double', Date='date')
        layer_defn = layer.GetLayerDefn()
        layer_infos = [(layer_defn.GetFieldDefn(i).GetName(),
            ESRITypes[layer_defn.GetFieldDefn(i).GetTypeName()]) for i in xrange(layer_defn.GetFieldCount())]
        
        # setup geometry field
        ret = plugin_lookout_fields.validate_and_insert(
                table_id = table_id,
                field_name = 'the_geom',
                field_label = 'Geometric feature',
                field_type = 'geometry'
            )
        if ret.errors:
            raise Exception(str(ret.errors))
        
        # setup attributes fields
        for field_name, field_type in layer_infos:
            ret = plugin_lookout_fields.validate_and_insert(
                table_id = table_id,
                field_name = field_name.lower(),
                field_label = field_name,
                field_type = field_type
            )
            if ret.errors:
                raise Exception(str(ret.errors))

def initFromFile(fileName, path, table_id, db, ext_table):
    filePath = os.path.join(path, fileName)
    fileExt = fileName.split('.')[-1]
    if fileExt in ('xls', 'xlsx', ):
        xlsx = load_workbook(filePath)
        first_sheet_name = xlsx.get_sheet_names()[0]
        sheet = xlsx.get_sheet_by_name(first_sheet_name)
        
        if ext_table._db(ext_table).count(): ext_table.drop()
        
        error = None
        for index,row in enumerate(sheet.rows[1:]):
            values = [cell.value for cell in row]
            fields = [r.field_name for r in db(db.plugin_lookout_fields.table_id==table_id).select(db.plugin_lookout_fields.field_name)]
            
            kwargs = dict([(k,v) for k,v in zip(fields, values)])
            ret = ext_table.validate_and_insert(**kwargs)
            if ret.errors:
                db.rollback()
                error = dict([(k, (kwargs[k], ret.errors[k])) for k in ret.errors])
                raise Exception(str(error))

    elif fileExt in ('zip', 'egg', 'jar', 'tar', 'gz', 'tgz', 'bz2', 'tz2', ):
        # uncompress the archive in new_dir
        new_dir = '.'.join(fileName.split('.')[:-1])
        new_path = os.path.join(path, new_dir)
        try:
            os.mkdir(new_path)
        except OSError, error:
            pass
        else:
            if fileExt in ('gz', 'bz2', ):
                filePath_new = '.'.join(filePath.split('.')[:-1] + ['tar', fileExt])
                os.rename(filePath, filePath_new)
                filePath = filePath_new
            extract(filePath, new_path)

        main_shp = [i for i in os.listdir(new_path) if i.split('.')[-1]=='shp'][0]
        shp_path = os.path.join(new_path, main_shp)
        
        driver = ogr.GetDriverByName('ESRI Shapefile')
        source = driver.Open(shp_path, 0)
        layer = source.GetLayer()
        
        # inspect field names and types
        ESRITypes = dict(String='string', Real='double', Date='date')
        layer_defn = layer.GetLayerDefn()
        layer_infos = [(layer_defn.GetFieldDefn(i).GetName(),
            ESRITypes[layer_defn.GetFieldDefn(i).GetTypeName()]) for i in xrange(layer_defn.GetFieldCount())]
        
        for index in xrange(layer.GetFeatureCount()):
            feature = layer.GetFeature(index)
            kwargs = dict([(fn[0].lower(), feature.GetField(fn[0])) for fn in layer_infos if feature.GetField(fn[0]) not in (None, '', '0000/00/00', )])
            if not hasattr(ext_table['the_geom'], 'st_asgeojson'):
                kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkt() # tested with postgis and web2py 1.99.7
            else:
                kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkb() # to be tested with web2py trunk with gis support
            
            ret = ext_table.validate_and_insert(**kwargs)
            if ret.errors:
                error = dict([(k, (kwargs[k], ret.errors[k])) for k in ret.errors])
                raise IOError(str(error))

#import ogr
#def get_layer(shp_path):
#    driver = ogr.GetDriverByName('ESRI Shapefile')
#    source = driver.Open(shp_path, 0)
#    return source.GetLayer()

#def get_layer_infos(layer):
#    '''
#    layer: OGR Layer object
#    layer_infos = [(<layer_name>, <layer_type>)]
#    '''
#    ESRITypes = dict(String='string', Real='double', Date='date')
#    layer_defn = layer.GetLayerDefn()
#    layer_infos = [(layer_defn.GetFieldDefn(i).GetName(),
#        ESRITypes[layer_defn.GetFieldDefn(i).GetTypeName()]) for i in xrange(layer_defn.GetFieldCount())]
#    return layer_infos

#def upload_layer(layer, table):
#    ids = []
#    layer_infos = get_layer_infos(layer)
#    for index in xrange(layer.GetFeatureCount()):
#        feature = layer.GetFeature(index)
#        kwargs = dict([(fn[0], feature.GetField(fn[0])) for fn in layer_infos])
#        if not hasattr(table['the_geom'], 'st_asgeojson'):
#            kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkb()
#        else:
#            kwargs['the_geom'] = feature.GetGeometryRef().ExportToWkt()
#        
#        ret = table.validate_and_insert(**kwargs)
#        
#        if ret.errors:
#            raise IOError(str(ret.errors))
#        else:
#            ids.append = ret.id
#    return tuple(ids)
            

from gluon import dal
def querysum(*args):
    db = args[0]._db
    args1 = args[len(args)%2:]
    r = map(db._adapter.AND, args1[::2], args1[1::2])
    while len(r) > 1:
        r = querysum(*r)
    if len(args)%2:
        r = map(db._adapter.AND, r, args[:len(args)%2])
    return dal.Query(db, *r)   
        

        
        
        
        
        
        
        
        
        
        
